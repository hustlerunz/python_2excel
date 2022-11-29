import openpyxl
import configparser
import sys
import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
  user="",
  password="",
  database="inventory_db"
)

mycursor = mydb.cursor()

config = configparser.ConfigParser()
config.read('config.ini')
filename_sheet = config['file']['file_name']
chkid_colume = config['check_id']['colume']
chkid_row = config['check_id']['row']
#result_colume = config['result']['rcolume']
#speed = config['request_persec']['cspeed']

book = openpyxl.load_workbook(filename_sheet)
sheet = book.active
row_count = sheet.max_row
print("row_count = ",row_count)
num_rec = 0
for x in range(row_count-2):
    row = x+int(chkid_row)
    site_id = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)).value
    inv_type = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+1).value
    inv_no = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+2).value
    inv_model = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+3).value
    inv_year = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+4).value
    inv_state = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+5).value
    inv_source = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+6).value
    inv_comment = sheet.cell(row=x+int(chkid_row), column=int(chkid_colume)+7).value
    print(site_id,inv_type,inv_no,inv_model,inv_year,inv_state,inv_source,inv_comment)
    if site_id == None:
            print("Out of data Or Success")
            
    else:
      sql = "INSERT INTO inventory (inv_code_area,inv_id,inv_type,inv_model,inv_buy_year,inv_status,inv_comment,inv_source) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
      val = (site_id,inv_no,inv_type,inv_model,inv_year,inv_state,inv_comment,inv_source)
      mycursor.execute(sql, val)
      mydb.commit()
      num_rec +=1
    #print(mycursor.rowcount, "record inserted.")
print("record num = ",num_rec)
book.save(filename_sheet)   
book.close()
sys.exit()
